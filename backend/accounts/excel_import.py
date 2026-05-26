"""
Generic Excel import view for admin use.
Supports: staff users, achievements, plots
"""
import openpyxl
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAdminUser
from django.contrib.auth import get_user_model
from achievements.models import Achievement
from projects.models import Project
from plots.models import Plot

User = get_user_model()


class ExcelImportView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdminUser]

    def post(self, request):
        file = request.FILES.get('file')
        import_type = request.data.get('type', 'staff')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower().replace(' ', '_') if c.value else '' for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(dict(zip(headers, row)))
        except Exception as e:
            return Response({'error': f'Could not read Excel file: {e}'}, status=400)

        if import_type == 'staff':
            return self._import_staff(rows)
        elif import_type == 'achievements':
            return self._import_achievements(rows)
        elif import_type == 'plots':
            project_id = request.data.get('project_id')
            return self._import_plots(rows, project_id)
        else:
            return Response({'error': f'Unknown import type: {import_type}'}, status=400)

    def _import_staff(self, rows):
        created, updated, errors = 0, 0, []
        for i, row in enumerate(rows, start=2):
            try:
                full_name = str(row.get('full_name') or row.get('name') or '').strip()
                if not full_name:
                    continue
                parts = full_name.split()
                first = parts[0] if parts else ''
                last = ' '.join(parts[1:]) if len(parts) > 1 else ''
                username = (first + last).lower().replace(' ', '')
                emp_id = str(row.get('employee_id') or row.get('emp_id') or '').strip()
                phone = str(row.get('phone') or row.get('mobile') or '').strip()
                position = str(row.get('position') or row.get('designation') or '').strip()
                department = str(row.get('department') or '').strip()
                site_location = str(row.get('site_location') or row.get('location') or '').strip()

                user, created_flag = User.objects.get_or_create(username=username)
                user.first_name = first
                user.last_name = last
                user.role = 'staff'
                user.is_staff = False
                if emp_id: user.employee_id = emp_id
                if phone: user.phone = phone
                if position: user.position = position
                if department: user.department = department
                if site_location: user.site_location = site_location
                if created_flag:
                    user.set_password('staff@2026')
                    created += 1
                else:
                    updated += 1
                user.save()
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        return Response({'created': created, 'updated': updated, 'errors': errors})

    def _import_achievements(self, rows):
        created, errors = 0, []
        for i, row in enumerate(rows, start=2):
            try:
                username = str(row.get('username') or '').strip()
                if not username:
                    continue
                user = User.objects.filter(username=username).first()
                if not user:
                    errors.append(f'Row {i}: User "{username}" not found')
                    continue
                Achievement.objects.create(
                    user=user,
                    period_type=str(row.get('period_type') or 'monthly').strip(),
                    period_label=str(row.get('period_label') or '').strip(),
                    plots_sold=int(row.get('plots_sold') or 0),
                    revenue=float(row.get('revenue') or 0),
                    rank=int(row.get('rank') or 0),
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        return Response({'created': created, 'errors': errors})

    def _import_plots(self, rows, project_id=None):
        created, updated, errors = 0, 0, []
        valid_statuses = {'available', 'booked', 'in_process', 'blocked', 'sold'}
        for i, row in enumerate(rows, start=2):
            try:
                # Resolve project
                proj_name = str(row.get('project_name') or '').strip()
                if project_id:
                    project = Project.objects.filter(id=project_id).first()
                elif proj_name:
                    project = Project.objects.filter(name__iexact=proj_name).first()
                else:
                    project = None
                if not project:
                    errors.append(f'Row {i}: Project "{proj_name}" not found')
                    continue

                plot_no = str(row.get('plot_no') or row.get('plot_number') or '').strip()
                if not plot_no:
                    errors.append(f'Row {i}: plot_no is required')
                    continue

                raw_status = str(row.get('status') or 'available').strip().lower()
                plot_status = raw_status if raw_status in valid_statuses else 'available'

                def to_float(val):
                    try: return float(str(val).replace(',', '').strip()) if val else None
                    except: return None

                area        = to_float(row.get('area_sqft') or row.get('area'))
                rate        = to_float(row.get('rate_per_sqft') or row.get('rate'))
                total_cost  = to_float(row.get('total_cost'))
                road_width  = str(row.get('road_width') or '').strip() or None
                survey_no   = str(row.get('survey_no') or '').strip() or None
                facing      = str(row.get('facing') or '').strip() or None

                # Auto-calc total cost if not provided
                if not total_cost and area and rate:
                    total_cost = area * rate

                plot, created_flag = Plot.objects.get_or_create(
                    project=project, plot_no=plot_no
                )
                if area is not None:       plot.area_sqft     = area
                if rate is not None:       plot.rate_per_sqft = rate
                if total_cost is not None: plot.total_cost    = total_cost
                if road_width:             plot.road_width    = road_width
                if survey_no:              plot.survey_no     = survey_no
                if facing:                 plot.facing        = facing
                plot.status = plot_status
                plot.save()
                if created_flag: created += 1
                else: updated += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        msg = f'Imported {created} new + {updated} updated plots.'
        if errors: msg += f' {len(errors)} row(s) skipped.'
        return Response({'message': msg, 'created': created, 'updated': updated, 'errors': errors})
