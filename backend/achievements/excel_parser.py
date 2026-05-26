"""
Parse achievement upload Excel files.

Supports two modes:
  A) Single-sheet  — month/year taken from the upload form.
  B) Multi-sheet   — sheet name is a month name (e.g. "March", "April", "May").
                     Year is always taken from the form. All sheets are processed.

Template columns (Sales):
  Employee Name | Designation | Team Name | Site Visits | Appointments |
  Meetings | Bookings | Registrations | Sq.Ft Sold | Units Sold

Template columns (Pre-Sales):
  Employee Name | Designation | Team Name | Site Visits | Appointments | Meetings

Identity key for upsert: (employee_name, team_name, team_type)
Every uploaded cell overwrites the stored value — full cell-level update.
"""
import openpyxl
from io import BytesIO

# Month-name → month-number
MONTH_MAP = {
    'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
    'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
    'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,
    'sep':9,'sept':9,'oct':10,'nov':11,'dec':12,
}

# Normalised header text → internal field name
HEADER_MAP = {
    'employee name':   'employee_name',
    'employee_name':   'employee_name',
    'name':            'employee_name',
    'emp name':        'employee_name',
    'designation':     'designation',
    'department':      'department',
    'dept':            'department',
    'team type':       'team_type',
    'team_type':       'team_type',
    'type':            'team_type',
    'team name':       'team_name',
    'team_name':       'team_name',
    'team':            'team_name',
    'site visit':      'site_visits',
    'site visits':     'site_visits',
    'site_visits':     'site_visits',
    'appt':            'appointments',
    'appointment':     'appointments',
    'appointments':    'appointments',
    'meeting':         'meetings',
    'meetings':        'meetings',
    'booking':         'bookings',
    'bookings':        'bookings',
    'reg':             'registrations',
    'registration':    'registrations',
    'registrations':   'registrations',
    'sq.ft':           'square_feet_sold',
    'sq ft':           'square_feet_sold',
    'sqft':            'square_feet_sold',
    'sq.ft sold':      'square_feet_sold',
    'square feet':     'square_feet_sold',
    'square feet sold':'square_feet_sold',
    'square_feet_sold':'square_feet_sold',
    'units':           'units_sold',
    'units sold':      'units_sold',
    'units_sold':      'units_sold',
}

REQUIRED = ['employee_name', 'team_name']
EMPTY_VALS = {'', '-', 'nil', 'n/a', 'na', 'none', 'null', 'total', 'overall', 'over all'}

# Rows whose employee_name starts with these are skipped (TOTAL / HEADER / NOTE rows)
SKIP_PREFIXES = ('total', 'over all', 'overall', 'pre sales', 'presales', 'sales team',
                 'pre-sales', 'grand total', '*fixed', '*required', '*')


def _to_int(v, default=0):
    if v is None: return default
    s = str(v).strip().lower()
    if s in EMPTY_VALS: return default
    try: return int(float(s))
    except Exception: return default


def _to_float(v, default=0.0):
    if v is None: return default
    s = str(v).strip().lower()
    if s in EMPTY_VALS: return default
    try: return float(s)
    except Exception: return default


def _norm_team_type(raw):
    s = str(raw or '').strip().lower().replace('-', '').replace(' ', '')
    if s.startswith('pre') or 'presales' in s:
        return 'pre_sales'
    if 'sales' in s:
        return 'sales'
    return None


def _month_from_sheet(name: str):
    return MONTH_MAP.get(name.strip().lower())


def _parse_sheet(ws, month, year, team_type_override=None):
    """
    Parse one worksheet and return (records, errors).
    team_type_override: if set, use this for every row (useful when sheet is
    a pre-labelled sales or pre-sales template).
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    # ── Locate header row ────────────────────────────────────────────────────
    header_idx = None
    for i, row in enumerate(all_rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) < 2:
            continue
        hits = sum(1 for c in non_empty if str(c).strip().lower() in HEADER_MAP)
        if hits >= 2:
            header_idx = i
            break

    if header_idx is None:
        return [], [{'row': 0, 'employee_name': '', 'team_name': '', 'errors': [
            f'No header row found in sheet "{ws.title}"'
        ]}]

    # ── Column map ───────────────────────────────────────────────────────────
    col_map = {}
    for j, h in enumerate(all_rows[header_idx]):
        if h is None: continue
        key = str(h).strip().lower()
        if key in HEADER_MAP:
            col_map[HEADER_MAP[key]] = j

    missing = [c for c in REQUIRED if c not in col_map]
    if missing:
        return [], [{'row': header_idx + 1, 'employee_name': '', 'team_name': '', 'errors': [
            f'Missing column(s) in sheet "{ws.title}": {", ".join(missing)}'
        ]}]

    records, errors = [], []

    for row_idx in range(header_idx + 1, len(all_rows)):
        row = all_rows[row_idx]
        if all(c is None or str(c).strip() == '' for c in row):
            continue   # blank row

        excel_row = row_idx + 1

        def get(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row): return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        emp_name = get('employee_name')
        team_nm  = get('team_name')

        # Skip TOTAL / section-header / template-note rows
        if emp_name and emp_name.lower().split()[0] in ('total', 'overall', 'over'):
            continue
        if emp_name and any(emp_name.lower().startswith(p) for p in SKIP_PREFIXES):
            continue
        if emp_name and emp_name.startswith('*'):
            continue

        row_errs = []
        if not emp_name or emp_name.lower() in EMPTY_VALS:
            row_errs.append('Employee Name is required')
        if not team_nm or team_nm.lower() in EMPTY_VALS:
            row_errs.append('Team Name is required')

        # Team type: use override if set (template-sourced), else read from column
        if team_type_override:
            team_type = team_type_override
        else:
            tt_raw    = get('team_type')
            team_type = _norm_team_type(tt_raw)
            if tt_raw and not team_type:
                row_errs.append(f'Unknown Team Type "{tt_raw}" — use Sales or Pre-Sales')
            elif not tt_raw or tt_raw.lower() in EMPTY_VALS:
                row_errs.append('Team Type is required (Sales or Pre-Sales)')

        if row_errs:
            errors.append({
                'row':           excel_row,
                'employee_name': emp_name or '',
                'team_name':     team_nm  or '',
                'errors':        row_errs,
            })
            continue

        records.append({
            'employee_name':    emp_name,
            'designation':      get('designation') or '',
            'department':       get('department')  or '',
            'team_type':        team_type,
            'team_name':        team_nm,
            'site_visits':      _to_int(get('site_visits')),
            'appointments':     _to_int(get('appointments')),
            'meetings':         _to_int(get('meetings')),
            'bookings':         _to_int(get('bookings')),
            'registrations':    _to_int(get('registrations')),
            'square_feet_sold': _to_float(get('square_feet_sold')),
            'units_sold':       _to_int(get('units_sold')),
            'month':            month,
            'year':             year,
        })

    return records, errors


def parse_achievements_excel(file_obj, default_month, default_year, team_type_override=None):
    """
    Parse file and return (all_records, all_errors).

    Multi-sheet: if a sheet name matches a month name, that month is used for
    that sheet's records (year always comes from the form).
    Single-sheet: default_month/year are used.
    """
    try:
        content = file_obj.read() if hasattr(file_obj, 'read') else file_obj
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        return [], [{'row': 0, 'employee_name': '', 'team_name': '',
                     'errors': [f'Cannot read Excel file: {exc}']}]

    all_records, all_errors = [], []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Detect month from sheet name, fall back to default
        detected_month = _month_from_sheet(sheet_name)
        month = detected_month if detected_month else default_month
        year  = default_year

        recs, errs = _parse_sheet(ws, month, year, team_type_override)
        all_records.extend(recs)
        all_errors.extend(errs)

    return all_records, all_errors
