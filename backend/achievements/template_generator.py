"""
Generate pre-populated Excel upload templates for Sales and Pre-Sales teams.
Team members and structure are taken from the master Excel file.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# ── Colour palette ───────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', start_color='1E3A5F')   # dark navy
HDR_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)

TEAM_FILL  = PatternFill('solid', start_color='2D6A4F')   # dark green  (Sales)
TEAM_FILL2 = PatternFill('solid', start_color='6C63FF')   # purple      (Pre-Sales)
TEAM_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=9)

HEAD_FILL  = PatternFill('solid', start_color='D4EDDA')   # light green
HEAD_FONT  = Font(bold=True, color='155724', name='Arial', size=9)

NOTE_FILL  = PatternFill('solid', start_color='FFF3CD')
NOTE_FONT  = Font(italic=True, color='856404', name='Arial', size=8)

BODY_FONT  = Font(name='Arial', size=9)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

thin = Side(style='thin', color='DDDDDD')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)


def _cell(ws, row, col, value='', font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ── Predefined team data (from master Excel) ─────────────────────────────────

SALES_TEAMS = [
    {
        'team_name': 'Jupila Madhesh',
        'members': [
            ('Jupila Madhesh',       'Project Head'),
            ('Samraj willson',       'BDM'),
            ('Prabakaran',           'Senior Sales Manager'),
            ('Rajakumara singh',     'Sales Manager'),
            ('Deliphin',             'Sales Manager'),
            ('Godwin',               'Sales Manager'),
        ],
    },
    {
        'team_name': 'Subash Chandra Bose G',
        'members': [
            ('Subash Chandra Bose G','Project Head'),
            ('Vignesh',              'BDM'),
            ('Srinivasan',           'BDM'),
            ('Jai Kumar',            'BDM'),
            ('Shibu',                'Sales Manager'),
            ('Sweth',                'Sales Manager'),
        ],
    },
    {
        'team_name': 'Ram Saravanan',
        'members': [
            ('Ram Saravanan',        'Project Head'),
            ('Yeswanth',             'BDM'),
            ('Sathish',              'Sales Manager'),
            ('Jagan Mohan Reddy',    'Sales Manager'),
            ('Siva Ram',             'Sales Manager'),
            ('Vignesh S',            'Sales Manager'),
        ],
    },
    {
        'team_name': 'Prabhu Durai',
        'members': [
            ('Prabhu Durai',         'Branch Head'),
            ('Yuvaraj',              'BDM'),
            ('Vikram D',             'Sales Manager'),
            ('Mohammed Asim',        'Sales Manager'),
            ('Manikandan M',         'Sales Manager'),
            ('Manikandan R',         'Sales Manager'),
            ('Devaraj S',            'Sales Manager'),
            ('Arun A',               'Sales Manager'),
            ('Gomathi Nayagam R',    'Sales Manager'),
        ],
    },
]

PRESALES_TEAMS = [
    {
        'team_name': 'Riya',
        'members': [
            ('Riya',            'Head'),
            ('Sudharhini D',    'Telecaller'),
            ('Jayaseeli',       'Telecaller'),
            ('Suguna',          'Telecaller'),
            ('Pooja',           'Telecaller'),
            ('Ester Rani',      'Telecaller'),
            ('Yamuna',          'Telecaller'),
            ('Haritha',         'Telecaller'),
            ('Pavithra',        'Telecaller'),
            ('Logeshwari',      'Telecaller'),
        ],
    },
    {
        'team_name': 'Prabhu Durai',
        'members': [
            ('Prabhu Durai',     'Head'),
            ('Priyadharshini',   'Telecaller'),
            ('Yogeshwari',       'Telecaller'),
            ('Supriya',          'Telecaller'),
            ('Keerthika',        'Telecaller'),
        ],
    },
]


# ── Sales template ────────────────────────────────────────────────────────────

SALES_HEADERS = [
    'Employee Name', 'Designation', 'Team Name',
    'Site Visits', 'Appointments', 'Meetings',
    'Bookings', 'Registrations', 'Sq.Ft Sold', 'Units Sold',
]
SALES_WIDTHS  = [25, 22, 25, 13, 15, 13, 13, 15, 14, 13]
SALES_NOTES   = [
    '*Required', 'e.g. Sales Manager', '*Fixed — do not change',
    'Count', 'Count', 'Count', 'Count', 'Count', 'Number', 'Count',
]
METRIC_START_COL = 4   # first metric column (D) — same for both templates

NUM_FILL = PatternFill('solid', start_color='F0FFF4')   # faint green for metric cells


def generate_sales_template() -> BytesIO:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Sales Team'

    # Column widths
    for col, w in enumerate(SALES_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    row = 1

    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SALES_HEADERS))
    c = ws.cell(row=row, column=1, value='SALES TEAM — MONTHLY ACHIEVEMENT DATA')
    c.font      = Font(bold=True, color='1E3A5F', name='Arial', size=11)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Row 2: note ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SALES_HEADERS))
    c = ws.cell(row=row, column=1,
                value='Fill in the metric columns (D–J) only. Do not edit Employee Name or Team Name.')
    c.font      = NOTE_FONT
    c.fill      = NOTE_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Row 3: column headers ─────────────────────────────────────────────────
    for col, hdr in enumerate(SALES_HEADERS, 1):
        _cell(ws, row, col, hdr, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)
    ws.row_dimensions[row].height = 22
    hdr_row = row
    row += 1

    ws.freeze_panes = f'A{row}'  # freeze header rows (no notes row — avoids parser confusion)

    # ── Team buckets ─────────────────────────────────────────────────────────
    for team in SALES_TEAMS:
        # Team header band
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SALES_HEADERS))
        c = ws.cell(row=row, column=1, value=f'  ▶  TEAM: {team["team_name"]}')
        c.font      = TEAM_FONT
        c.fill      = TEAM_FILL
        c.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        for (name, desig) in team['members']:
            # Name
            _cell(ws, row, 1, name,           font=BODY_FONT, align=LEFT,   border=BORDER)
            # Designation
            _cell(ws, row, 2, desig,          font=BODY_FONT, align=LEFT,   border=BORDER)
            # Team name (locked display)
            _cell(ws, row, 3, team['team_name'], font=BODY_FONT, align=LEFT, border=BORDER)
            # Metric cells (empty, light fill)
            for col in range(4, len(SALES_HEADERS) + 1):
                _cell(ws, row, col, 0, font=BODY_FONT, fill=NUM_FILL, align=RIGHT, border=BORDER)
            ws.row_dimensions[row].height = 16
            row += 1

        # Team total formula row
        start = row - len(team['members'])
        end   = row - 1
        _cell(ws, row, 1, f'TOTAL — {team["team_name"]}',
              font=Font(bold=True, name='Arial', size=9), align=LEFT, border=BORDER)
        _cell(ws, row, 2, '', border=BORDER)
        _cell(ws, row, 3, '', border=BORDER)
        for col in range(4, len(SALES_HEADERS) + 1):
            letter = openpyxl.utils.get_column_letter(col)
            c = ws.cell(row=row, column=col,
                        value=f'=SUM({letter}{start}:{letter}{end})')
            c.font      = Font(bold=True, name='Arial', size=9)
            c.alignment = RIGHT
            c.border    = BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        row += 1   # blank separator

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Pre-Sales template ────────────────────────────────────────────────────────

PRESALES_HEADERS = [
    'Employee Name', 'Designation', 'Team Name',
    'Site Visits', 'Appointments', 'Meetings',
]
PRESALES_WIDTHS = [25, 22, 20, 13, 15, 13]
PRESALES_NOTES  = [
    '*Required', 'e.g. Telecaller', '*Fixed — do not change',
    'Count', 'Count', 'Count',
]


def generate_presales_template() -> BytesIO:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Pre-Sales Team'

    for col, w in enumerate(PRESALES_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(PRESALES_HEADERS))
    c = ws.cell(row=row, column=1, value='PRE-SALES TEAM — MONTHLY ACHIEVEMENT DATA')
    c.font      = Font(bold=True, color='6C63FF', name='Arial', size=11)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(PRESALES_HEADERS))
    c = ws.cell(row=row, column=1,
                value='Fill in the metric columns (D–F) only. Do not edit Employee Name or Team Name.')
    c.font      = NOTE_FONT
    c.fill      = NOTE_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 16
    row += 1

    for col, hdr in enumerate(PRESALES_HEADERS, 1):
        _cell(ws, row, col, hdr, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)
    ws.row_dimensions[row].height = 22
    row += 1

    ws.freeze_panes = f'A{row}'

    NUM_FILL2 = PatternFill('solid', start_color='F3F0FF')   # faint purple for pre-sales

    for team in PRESALES_TEAMS:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(PRESALES_HEADERS))
        c = ws.cell(row=row, column=1, value=f'  ▶  TEAM: {team["team_name"]}')
        c.font      = TEAM_FONT
        c.fill      = TEAM_FILL2
        c.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        for (name, desig) in team['members']:
            _cell(ws, row, 1, name,              font=BODY_FONT, align=LEFT,  border=BORDER)
            _cell(ws, row, 2, desig,             font=BODY_FONT, align=LEFT,  border=BORDER)
            _cell(ws, row, 3, team['team_name'], font=BODY_FONT, align=LEFT,  border=BORDER)
            for col in range(4, len(PRESALES_HEADERS) + 1):
                _cell(ws, row, col, 0, font=BODY_FONT, fill=NUM_FILL2, align=RIGHT, border=BORDER)
            ws.row_dimensions[row].height = 16
            row += 1

        # Total formula
        start = row - len(team['members'])
        end   = row - 1
        _cell(ws, row, 1, f'TOTAL — {team["team_name"]}',
              font=Font(bold=True, name='Arial', size=9), align=LEFT, border=BORDER)
        _cell(ws, row, 2, '', border=BORDER)
        _cell(ws, row, 3, '', border=BORDER)
        for col in range(4, len(PRESALES_HEADERS) + 1):
            letter = openpyxl.utils.get_column_letter(col)
            c = ws.cell(row=row, column=col,
                        value=f'=SUM({letter}{start}:{letter}{end})')
            c.font      = Font(bold=True, name='Arial', size=9)
            c.alignment = RIGHT
            c.border    = BORDER
        ws.row_dimensions[row].height = 16
        row += 1
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
