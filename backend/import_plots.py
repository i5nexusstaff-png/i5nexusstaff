"""
Import all plot data from the Excel file into the database.
Run with: py -3 import_plots.py
"""

import os, sys, django

sys.path.insert(0, r'D:\i5nexus\backend')
os.environ['DJANGO_SETTINGS_MODULE'] = 'config.settings'
django.setup()

import openpyxl
from projects.models import Project
from plots.models import Plot

FILE = r'C:\Users\georg\Downloads\Plots sold and unsold (1).xlsx'
wb   = openpyxl.load_workbook(FILE, data_only=True)

# ── helpers ───────────────────────────────────────────────────────────────────
def cv(v):
    """Return a clean string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ('none', 'nan', '-', 'n/a', 'null', 'nil', 'na') else None

def fl(v):
    """Parse a number from a cell value (handles ₹ signs and commas)."""
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '').replace('₹', '').replace(' ', '').strip())
    except Exception:
        return None

HEADER_WORDS = {
    'plot', 'sqft', 'sq.ft', 'sq ft', 'rate', 'facing', 'road', 'width',
    'area', 'sft', 'sno', 's.no', 'sold', 'available', 'survey', 'phase',
    'section', 'yet', 'total', 'cost', 'no.', 'plots', 'sl.no', 'serial',
    'description', 'name', 'type', 'yet to', 'const', 'remarks',
}

def is_header(v):
    """Return True if a cell value looks like a column header, not a plot number."""
    if v is None:
        return True
    s = str(v).strip().lower()
    if not s:
        return True
    for w in HEADER_WORDS:
        if w in s:
            return True
    return False

def is_valid_plot_no(v):
    """Return True if this is a plausible plot number (has at least one digit or is ≤3 chars)."""
    if v is None:
        return False
    s = str(v).strip()
    if not s or len(s) > 25:
        return False
    has_digit = any(c.isdigit() for c in s)
    # Allow short alphabetic codes (A, B, C) but not long words
    return has_digit or (len(s) <= 3 and s.isalpha())

def get_project(name):
    p, created = Project.objects.get_or_create(
        name=name, defaults={'location': '', 'description': ''}
    )
    status = '[+] Created' if created else '[=] Found'
    print(f"  {status}: {name}")
    return p

def upsert(project, plot_no, area=None, facing=None, rate=None,
           total_cost=None, road_width=None, survey_no=None,
           status='available', notes=''):
    """Create or update a plot. Sold status is never downgraded."""
    if not is_valid_plot_no(plot_no):
        return None
    pno = str(plot_no).strip()

    defaults = {'status': status}
    if area is not None:       defaults['area_sqft']     = area
    if facing:                 defaults['facing']         = facing.strip()[:100]
    if rate is not None:       defaults['rate_per_sqft']  = rate
    if total_cost is not None: defaults['total_cost']     = total_cost
    if road_width:             defaults['road_width']     = str(road_width).strip()[:100]
    if survey_no:              defaults['survey_no']      = str(survey_no).strip()[:100]
    if notes:                  defaults['notes']          = notes.strip()

    try:
        plot, created = Plot.objects.get_or_create(
            project=project, plot_no=pno, defaults=defaults
        )
        if not created:
            changed = False
            if area       and not plot.area_sqft:     plot.area_sqft     = area;       changed = True
            if facing     and not plot.facing:         plot.facing        = facing.strip()[:100]; changed = True
            if rate       and not plot.rate_per_sqft:  plot.rate_per_sqft = rate;       changed = True
            if total_cost and not plot.total_cost:     plot.total_cost    = total_cost; changed = True
            if road_width and not plot.road_width:     plot.road_width    = str(road_width)[:100]; changed = True
            if survey_no  and not plot.survey_no:      plot.survey_no     = str(survey_no)[:100]; changed = True
            if notes      and not plot.notes:          plot.notes         = notes;      changed = True
            # Never downgrade sold → available/other
            if status == 'sold' and plot.status != 'sold':
                plot.status = 'sold'; changed = True
            elif status == 'available' and plot.status not in ('sold', 'booked', 'in_process'):
                plot.status = 'available'; changed = True
            if changed:
                plot.save()
        return plot
    except Exception as e:
        print(f"    [!] Plot {pno}: {e}")
        return None

STATS = {}

def tally(name, status):
    if name not in STATS:
        STATS[name] = {'sold': 0, 'available': 0}
    key = 'sold' if status == 'sold' else 'available'
    STATS[name][key] += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — i5 Palace City
# Layout: S.NO | PLOT_NO | SQFT | RATE | FACING  (repeated in 3 horizontal groups)
# Sold   → cols A(0)–E(4)  → plot_no=B(1), sqft=C(2), rate=D(3), facing=E(4)
# Avail1 → cols G(6)–K(10) → plot_no=H(7), sqft=I(8), rate=J(9), facing=K(10)
# Avail2 → cols L(11)–P(15)→ plot_no=M(12), sqft=N(13), rate=O(14), facing=P(15)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== i5 Palace City ===")
ws = wb['i5 Palace City']
pr = get_project('i5 Palace City')

for r in ws.iter_rows(min_row=4, max_row=60, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    # Sold
    pno = cv(_r(1))
    if pno and is_valid_plot_no(pno) and not is_header(pno):
        if upsert(pr, pno, fl(_r(2)), cv(_r(4)), fl(_r(3)), status='sold'):
            tally('i5 Palace City', 'sold')

    # Available set 1
    pno = cv(_r(7))
    if pno and is_valid_plot_no(pno) and not is_header(pno):
        if upsert(pr, pno, fl(_r(8)), cv(_r(10)), fl(_r(9)), status='available'):
            tally('i5 Palace City', 'available')

    # Available set 2
    pno = cv(_r(12))
    if pno and is_valid_plot_no(pno) and not is_header(pno):
        if upsert(pr, pno, fl(_r(13)), cv(_r(15)), fl(_r(14)), status='available'):
            tally('i5 Palace City', 'available')


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Aurowin Enclave
# Layout: PLOT_NO | AREA | FACING | ROAD_WIDTH | RATE | TOTAL_COST (6-col groups)
# Sold   → base col 0  (A–F)
# Avail1 → base col 7  (H–M)
# Avail2 → base col 13 (N–S)
# Avail3 → base col 19 (T–Y)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== Aurowin Enclave ===")
ws = wb['Aurowin Enclave']
pr = get_project('Aurowin Enclave')

SECTIONS_AE = [(0, 'sold'), (7, 'available'), (13, 'available'), (19, 'available')]

for r in ws.iter_rows(min_row=3, max_row=45, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    for base, status in SECTIONS_AE:
        pno = cv(_r(base))
        if not pno or not is_valid_plot_no(pno) or is_header(pno):
            continue
        area  = fl(_r(base + 1))
        fac   = cv(_r(base + 2))
        road  = cv(_r(base + 3))
        rate  = fl(_r(base + 4))
        tcost = fl(_r(base + 5))
        if upsert(pr, pno, area, fac, rate, tcost, road, status=status):
            tally('Aurowin Enclave', status)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — i5 WonderCity
# Sold:  plot numbers in columns A(0), B(1), C(2), D(3) — minimal/no area
# Avail: F(5)=plot_no, G(6)=sqft  and  H(7)=plot_no, I(8)=sqft
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== i5 WonderCity ===")
ws = wb['i5 WonderCity']
pr = get_project('i5 WonderCity')

for r in ws.iter_rows(min_row=3, max_row=25, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    # Sold plot numbers (no area) — columns A–D
    for col in range(4):
        pno = cv(_r(col))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            # Next cell might be area (float) or another plot no
            nxt = fl(_r(col + 1)) if col < 3 else None
            area = nxt if (nxt and nxt > 50) else None
            if upsert(pr, pno, area, status='sold'):
                tally('i5 WonderCity', 'sold')

    # Available: pairs at (5,6) and (7,8)
    for base in [5, 7]:
        pno  = cv(_r(base))
        area = fl(_r(base + 1))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            if upsert(pr, pno, area, status='available'):
                tally('i5 WonderCity', 'available')


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — i5 Global City
# Sold  (3 groups of 3): A-C, D-F, G-I  → plot_no | facing | sqft
# Avail (2 groups of 4): L-O, P-S        → plot_no | facing | sqft | rate
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== i5 Global City ===")
ws = wb['i5 Global City']
pr = get_project('i5 Global City')

for r in ws.iter_rows(min_row=3, max_row=170, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    for base in [0, 3, 6]:             # Sold groups
        pno  = cv(_r(base))
        fac  = cv(_r(base + 1))
        area = fl(_r(base + 2))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            if upsert(pr, pno, area, fac, status='sold'):
                tally('i5 Global City', 'sold')

    for base in [11, 15]:              # Available groups
        pno  = cv(_r(base))
        fac  = cv(_r(base + 1))
        area = fl(_r(base + 2))
        rate = fl(_r(base + 3))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            if upsert(pr, pno, area, fac, rate, status='available'):
                tally('i5 Global City', 'available')


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Sri Sai City
# Each row has the phase label directly as a Roman numeral in col A (sold) / D (avail)
# Sold  : A(0)=roman_phase, B(1)=plot_no, C(2)=area  (phases I, II, III)
# Avail : D(3)=roman_phase, E(4)=plot_no, F(5)=area  (phases IV, V, VI)
# Extra : H(7)=plot_no, I(8)=rate, J(9)=area         (Phase VI yet-to-sell)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== Sri Sai City ===")
ws   = wb['Sri Sai City']
pr   = get_project('Sri Sai City')

# Delete stale/bad plots and reimport clean
from plots.models import Plot as _Plot
_Plot.objects.filter(project=pr).delete()

ROMAN = {'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII'}
last_sold_phase  = ''
last_avail_phase = ''

for r in ws.iter_rows(min_row=4, max_row=100, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    # Phase label is the Roman numeral directly in col A / D
    ph_raw_s = cv(_r(0))
    if ph_raw_s and ph_raw_s.strip().upper() in ROMAN:
        last_sold_phase = ph_raw_s.strip().upper()

    ph_raw_a = cv(_r(3))
    if ph_raw_a and ph_raw_a.strip().upper() in ROMAN:
        last_avail_phase = ph_raw_a.strip().upper()

    # Sold: col B=plot_no, col C=area
    pno  = cv(_r(1))
    area = fl(_r(2))
    if pno and is_valid_plot_no(pno) and not is_header(pno) and last_sold_phase:
        full_pno = f"Ph{last_sold_phase}-{pno}"
        if upsert(pr, full_pno, area, notes=f'Phase {last_sold_phase}', status='sold'):
            tally('Sri Sai City', 'sold')

    # Available: col E=plot_no, col F=area
    pno  = cv(_r(4))
    area = fl(_r(5))
    if pno and is_valid_plot_no(pno) and not is_header(pno) and last_avail_phase:
        full_pno = f"Ph{last_avail_phase}-{pno}"
        if upsert(pr, full_pno, area, notes=f'Phase {last_avail_phase}', status='available'):
            tally('Sri Sai City', 'available')

    # Phase VI yet-to-sell: H(7)=plot_no, I(8)=rate, J(9)=area
    pno  = cv(_r(7))
    rate = fl(_r(8))
    area = fl(_r(9))
    if pno and is_valid_plot_no(pno) and not is_header(pno):
        if upsert(pr, str(pno), area, rate=rate, notes='Phase VI - Yet to sell', status='available'):
            tally('Sri Sai City', 'available')


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — OMR i5 City
# Sold1 : A(0)=plot_no, B(1)=facing, C(2)=sqft
# Sold2 : D(3)=plot_no, E(4)=facing, F(5)=sqft
# Avail : H(7)=plot_no, I(8)=facing, J(9)=sqft
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== OMR i5 City ===")
ws = wb['OMR i5 City']
pr = get_project('OMR i5 City')

for r in ws.iter_rows(min_row=3, max_row=60, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    for base, status in [(0, 'sold'), (3, 'sold'), (7, 'available')]:
        pno  = cv(_r(base))
        fac  = cv(_r(base + 1))
        area = fl(_r(base + 2))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            if upsert(pr, str(pno), area, fac, status=status):
                tally('OMR i5 City', status)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — CK Madhavan Nagar
# Sold  (5 groups of 3): A-C, D-F, G-I, J-L, M-O  → plot_no | area | facing
# Avail (1 group of 3) : Q(16)-S(18)               → plot_no | area | facing
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== CK Madhavan Nagar ===")
ws = wb['CK Madhavan Nagar']
pr = get_project('CK Madhavan Nagar')

for r in ws.iter_rows(min_row=3, max_row=65, values_only=True):
    def _r(i): return r[i] if len(r) > i else None

    for base in [0, 3, 6, 9, 12]:     # Sold groups
        pno  = cv(_r(base))
        area = fl(_r(base + 1))
        fac  = cv(_r(base + 2))
        if pno and is_valid_plot_no(pno) and not is_header(pno):
            if upsert(pr, str(pno), area, fac, status='sold'):
                tally('CK Madhavan Nagar', 'sold')

    # Available
    pno  = cv(_r(16))
    area = fl(_r(17))
    fac  = cv(_r(18))
    if pno and is_valid_plot_no(pno) and not is_header(pno):
        if upsert(pr, str(pno), area, fac, status='available'):
            tally('CK Madhavan Nagar', 'available')


# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 55)
print("  IMPORT SUMMARY")
print("=" * 55)
grand_sold = grand_avail = 0
for name, s in STATS.items():
    print(f"\n  {name}")
    print(f"    Sold:      {s['sold']:>4}  plots")
    print(f"    Available: {s['available']:>4}  plots")
    grand_sold  += s['sold']
    grand_avail += s['available']

print(f"\n  {'-'*50}")
print(f"  TOTAL  {grand_sold + grand_avail} plots   "
      f"({grand_sold} sold  +  {grand_avail} available)")
print("=" * 55)

# Update project total_plots counts
print("\n  Refreshing project totals…")
for p in Project.objects.all():
    count = p.plots.count()
    if p.total_plots != count:
        p.total_plots = count
        p.save(update_fields=['total_plots'])
        print(f"    {p.name}: {count} total plots")

print("\n  [done]\n")
