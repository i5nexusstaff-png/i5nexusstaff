import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.db.models import Count, Q
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from accounts.permissions import IsAdminRole

from .models import Project
from .serializers import ProjectListSerializer, ProjectDetailSerializer
from plots.models import Plot
from plots.serializers import PlotSerializer


# ── Status normaliser ─────────────────────────────────────────────────────────
STATUS_VALID = {'available', 'booked', 'in_process', 'blocked', 'sold'}
STATUS_ALIAS = {
    'available':  'available',
    'booked':     'booked',
    'in process': 'in_process',
    'in_process': 'in_process',
    'inprocess':  'in_process',
    'blocked':    'blocked',
    'sold':       'sold',
}

# ── Column name normaliser ────────────────────────────────────────────────────
# None = skip this column (e.g. S.No — reference only, not stored)
COL_MAP = {
    # S.No — skip
    's.no':            None,
    's. no':           None,
    'sno':             None,
    'serial no':       None,
    'serial_no':       None,
    's no':            None,
    # Plot No
    'plot no':         'plot_no',
    'plot_no':         'plot_no',
    'plot number':     'plot_no',
    'plotno':          'plot_no',
    # Area
    'area (sq.ft)':    'area_sqft',
    'area (sqft)':     'area_sqft',
    'area_sqft':       'area_sqft',
    'area':            'area_sqft',
    'area sq ft':      'area_sqft',
    # Facing
    'facing':          'facing',
    # Rate
    'rate/sq.ft':      'rate_per_sqft',
    'rate/sqft':       'rate_per_sqft',
    'rate per sq.ft':  'rate_per_sqft',
    'rate per sqft':   'rate_per_sqft',
    'rate_per_sqft':   'rate_per_sqft',
    'rate':            'rate_per_sqft',
    # Total Cost
    'total cost':      'total_cost',
    'total_cost':      'total_cost',
    'totalcost':       'total_cost',
    # Status
    'status':          'status',
}


def _safe_float(val):
    try:
        import math
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return None
        return float(val)
    except Exception:
        return None


def _safe_str(val):
    try:
        import math
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return ''
        return str(val).strip()
    except Exception:
        return ''


class ProjectViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Single query that annotates per-project plot counts.
        Eliminates the N+1 caused by the old @property methods — for 10 projects
        this was 30 extra COUNT queries; now it's 1 JOIN.
        """
        return Project.objects.annotate(
            sold_plots=Count('plots', filter=Q(plots__status='sold')),
            available_plots=Count('plots', filter=Q(plots__status='available')),
            total_plot_count=Count('plots'),
        ).order_by('name')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ProjectDetailSerializer
        return ProjectListSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy',
                           'import_plots', 'clear_plots', 'upload_layout']:
            return [IsAdminRole()]
        return super().get_permissions()

    def partial_update(self, request, *args, **kwargs):
        # ── Remove layout ──────────────────────────────────────────────────────
        if request.data.get('remove_layout') in (True, 'true', '1'):
            project = self.get_object()
            if project.layout_image:
                project.layout_image.delete(save=False)
                project.layout_image = None
                project.save(update_fields=['layout_image'])
            return Response(ProjectDetailSerializer(project, context={'request': request}).data)

        # ── 5 MB file size guards ──────────────────────────────────────────────
        MB5 = 5 * 1024 * 1024
        layout_file = request.FILES.get('layout_image')
        logo_file   = request.FILES.get('image')
        if layout_file and layout_file.size > MB5:
            return Response(
                {'error': 'Layout image must be under 5 MB. Please compress the image and try again.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if logo_file and logo_file.size > MB5:
            return Response(
                {'error': 'Logo image must be under 5 MB. Please compress the image and try again.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().partial_update(request, *args, **kwargs)

    # ── GET /projects/{id}/plots/ ─────────────────────────────────────────────
    @action(detail=True, methods=['get'])
    def plots(self, request, pk=None):
        project = self.get_object()
        plots = project.plots.all()
        status_filter = request.query_params.get('status')
        if status_filter:
            plots = plots.filter(status=status_filter)
        return Response(PlotSerializer(plots, many=True).data)

    # ── GET /projects/dashboard_stats/ ───────────────────────────────────────
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        # Single aggregate query for global counters
        agg = Plot.objects.aggregate(
            total_plots=Count('id'),
            total_sold=Count('id', filter=Q(status='sold')),
            total_available=Count('id', filter=Q(status='available')),
        )

        # Single annotated query for per-project breakdown (replaces Python loop)
        projects_qs = Project.objects.annotate(
            sold=Count('plots', filter=Q(plots__status='sold')),
            available=Count('plots', filter=Q(plots__status='available')),
            total=Count('plots'),
        ).order_by('name')

        project_stats = [
            {
                'id': p.id, 'name': p.name, 'location': p.location,
                'sold': p.sold, 'available': p.available, 'total': p.total,
                'sold_percentage': round((p.sold / p.total * 100), 1) if p.total > 0 else 0,
            }
            for p in projects_qs
        ]

        total = agg['total_plots'] or 0
        sold  = agg['total_sold']  or 0

        return Response({
            'total_projects': projects_qs.count(),
            'total_plots':    total,
            'total_sold':     sold,
            'total_available': agg['total_available'] or 0,
            'sold_percentage': round((sold / total * 100), 1) if total > 0 else 0,
            'project_stats':  project_stats,
        })

    # ── GET /projects/plot_template/ — download Excel template ───────────────
    @action(detail=False, methods=['get'], url_path='plot_template')
    def plot_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plots'

        # ── Styles ──
        hdr_fill   = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        data_font  = Font(name='Arial', size=10)
        bold_font  = Font(name='Arial', bold=True, size=10)
        alt_fill   = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        thin       = Side(style='thin', color='D1D5DB')
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_al  = Alignment(horizontal='center', vertical='center')
        left_al    = Alignment(horizontal='left',   vertical='center')

        # S.No column gets a lighter grey header to mark it as reference-only
        sno_fill = PatternFill(start_color='4B5563', end_color='4B5563', fill_type='solid')

        # ── Headers: S.No | Plot No | Facing | Area (sq.ft) | Rate per sq.ft | Total Cost | Status ──
        headers = [
            ('S.No',            6,  sno_fill),
            ('Plot No',         14, hdr_fill),
            ('Facing',          14, hdr_fill),
            ('Area (sq.ft)',    15, hdr_fill),
            ('Rate per sq.ft',  16, hdr_fill),
            ('Total Cost',      16, hdr_fill),
            ('Status',          16, hdr_fill),
        ]
        for col, (hdr, width, fill) in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=hdr)
            c.font = hdr_font; c.fill = fill; c.border = border
            c.alignment = center_al
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 22

        # ── Data validation for Status column (G) ──
        dv = DataValidation(
            type='list',
            formula1='"available,booked,in process,blocked,sold"',
            allow_blank=True, showDropDown=False,
            showErrorMessage=True,
            errorTitle='Invalid Status',
            error='Choose: available | booked | in process | blocked | sold',
        )
        dv.sqref = 'G2:G50000'
        ws.add_data_validation(dv)

        # ── Footer note (row 3 — just below the blank data row 2) ──
        note_row = 3
        note = ws.cell(row=note_row, column=1,
                       value='Status values: available | booked | in process | blocked | sold')
        note.font = Font(name='Arial', italic=True, color='6B7280', size=9)

        sub = ws.cell(row=note_row + 1, column=1,
                      value='S.No is for reference only and is ignored during import. '
                            'Total Cost is auto-calculated (Area × Rate) if left empty.')
        sub.font = Font(name='Arial', italic=True, color='9CA3AF', size=9)

        # ── Freeze header row ──
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="plots_template.xlsx"'
        return resp

    # ── POST /projects/{id}/import_plots/ — upload Excel, replace all plots ──
    @action(detail=True, methods=['post'], url_path='import_plots')
    def import_plots(self, request, pk=None):
        import math
        project = self.get_object()
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Parse file ──
        try:
            name = file.name.lower()
            if name.endswith('.csv'):
                import pandas as pd
                df = pd.read_csv(file)
            else:
                import pandas as pd
                df = pd.read_excel(file)
        except Exception as e:
            return Response({'error': f'Cannot read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Normalise columns (None = skip, e.g. S.No) ──
        df.columns = [COL_MAP.get(str(c).strip().lower(), str(c).strip().lower()) for c in df.columns]
        # Drop columns that were explicitly mapped to None (reference-only, e.g. S.No)
        df = df.loc[:, df.columns.notna()]

        if 'plot_no' not in df.columns:
            return Response(
                {'error': 'Missing required column "Plot No". Download the template for the correct format.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Delete existing plots ──
        project.plots.all().delete()

        # ── Build plot objects ──
        new_plots = []
        skipped   = 0

        for _, row in df.iterrows():
            plot_no = _safe_str(row.get('plot_no', ''))
            if not plot_no or plot_no.lower() == 'nan':
                skipped += 1
                continue

            area = _safe_float(row.get('area_sqft'))
            rate = _safe_float(row.get('rate_per_sqft'))
            cost = _safe_float(row.get('total_cost'))

            # Auto-calc total cost
            if cost is None and area is not None and rate is not None:
                cost = area * rate

            facing = _safe_str(row.get('facing', ''))

            raw_st = _safe_str(row.get('status', 'available')).lower()
            st = STATUS_ALIAS.get(raw_st, raw_st)
            if st not in STATUS_VALID:
                st = 'available'

            new_plots.append(Plot(
                project=project,
                plot_no=plot_no,
                area_sqft=area,
                facing=facing,
                rate_per_sqft=rate,
                total_cost=cost,
                status=st,
            ))

        # ── Bulk insert ──
        Plot.objects.bulk_create(new_plots, ignore_conflicts=True)

        # ── Update project total_plots count ──
        project.total_plots = project.plots.count()
        project.save(update_fields=['total_plots'])

        return Response({
            'message': f'Imported {len(new_plots)} plots into "{project.name}".'
                       + (f' ({skipped} empty rows skipped)' if skipped else ''),
            'count':   len(new_plots),
            'skipped': skipped,
        })

    # ── DELETE /projects/{id}/clear_plots/ — wipe all plots ─────────────────
    @action(detail=True, methods=['delete'], url_path='clear_plots')
    def clear_plots(self, request, pk=None):
        project = self.get_object()
        deleted, _ = project.plots.all().delete()
        project.total_plots = 0
        project.save(update_fields=['total_plots'])
        return Response({'message': f'Deleted {deleted} plots from "{project.name}".'})
