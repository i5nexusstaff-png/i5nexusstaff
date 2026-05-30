import io
import math
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import calendar

from rest_framework import viewsets, status as http_status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.http import HttpResponse

from .models import DailyReport, REPORT_SCHEMA, LEGACY_TYPE_MAP
from .serializers import DailyReportSerializer


def _notify_report_submitted(staff_user, report_date):
    try:
        from accounts.models import User
        from notifications.push_utils import create_and_push
        name  = staff_user.get_full_name() or staff_user.username
        title = '📄 Report Submitted'
        msg   = f'{name} submitted their daily report for {report_date}.'
        for admin in User.objects.filter(role__in=['admin', 'super_admin'], is_active=True):
            url = '/superadmin/reports' if admin.role == 'super_admin' else '/admin/reports'
            create_and_push(admin, title, msg, 'report', url=url)
    except Exception:
        pass


# ── Excel style helpers ────────────────────────────────────────────────────────
def _hdr_style():
    thin = Side(style='thin', color='B0BAC8')
    return {
        'font':  Font(name='Calibri', bold=True, color='FFFFFF', size=11),
        'fill':  PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid'),
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'align': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }

def _cell_style(alt=False):
    thin = Side(style='thin', color='D1D5DB')
    return {
        'font':  Font(name='Calibri', size=10),
        'fill':  PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid') if alt else None,
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'align': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }

def _apply(cell, style):
    cell.font   = style['font']
    if style.get('fill'): cell.fill = style['fill']
    cell.border = style['border']
    cell.alignment = style['align']


class DailyReportViewSet(viewsets.ModelViewSet):
    serializer_class   = DailyReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs   = DailyReport.objects.select_related('user').all()

        if user.role == 'staff':
            qs = qs.filter(user=user)
        elif user.role == 'admin':
            qs = qs.filter(user__role__in=['staff', 'admin'])

        p = self.request.query_params
        if p.get('date'):
            qs = qs.filter(report_date=p['date'])
        if p.get('week_start'):
            try:
                ws = date.fromisoformat(p['week_start'])
                qs = qs.filter(report_date__range=[ws, ws + timedelta(days=6)])
            except ValueError:
                pass
        if p.get('month'):
            try:
                y, m = p['month'].split('-')
                qs   = qs.filter(report_date__year=y, report_date__month=m)
            except ValueError:
                pass
        if p.get('user_id') and user.role in ('admin', 'super_admin'):
            qs = qs.filter(user_id=p['user_id'])
        if p.get('report_type'):
            qs = qs.filter(report_type=p['report_type'])
        if p.get('status'):
            qs = qs.filter(status=p['status'])

        return qs

    def perform_create(self, serializer):
        extra = {}
        if serializer.validated_data.get('status') == 'submitted':
            extra['submitted_at'] = timezone.now()
        serializer.save(user=self.request.user, **extra)

    def create(self, request, *args, **kwargs):
        user        = request.user
        report_date = request.data.get('report_date', timezone.localdate().isoformat())
        report_type = LEGACY_TYPE_MAP.get(
            request.data.get('report_type', ''),
            request.data.get('report_type', ''),
        )
        new_status  = request.data.get('status', 'draft')
        try:
            existing = DailyReport.objects.get(user=user, report_date=report_date, report_type=report_type)
            ser = self.get_serializer(existing, data={**request.data, 'report_type': report_type}, partial=True)
            ser.is_valid(raise_exception=True)
            extra = {}
            first_submit = new_status == 'submitted' and not existing.submitted_at
            if first_submit:
                extra['submitted_at'] = timezone.now()
            ser.save(**extra)
            if first_submit:
                _notify_report_submitted(user, report_date)
            return Response(ser.data)
        except DailyReport.DoesNotExist:
            data = {**request.data, 'report_type': report_type}
            ser = self.get_serializer(data=data)
            ser.is_valid(raise_exception=True)
            extra = {}
            if new_status == 'submitted':
                extra['submitted_at'] = timezone.now()
            ser.save(user=user, **extra)
            if new_status == 'submitted':
                _notify_report_submitted(user, report_date)
            return Response(ser.data, status=201)

    def update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        instance   = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        extra = {}
        if request.data.get('status') == 'submitted' and not instance.submitted_at:
            extra['submitted_at'] = timezone.now()
        serializer.save(**extra)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='schema')
    def schema(self, request):
        return Response(REPORT_SCHEMA)

    @action(detail=True, methods=['patch'], url_path='review')
    def review(self, request, pk=None):
        if request.user.role not in ('admin', 'super_admin'):
            return Response({'detail': 'Forbidden'}, status=403)
        report = self.get_object()
        if 'admin_notes' in request.data:
            report.admin_notes = request.data['admin_notes']
        if 'status' in request.data:
            report.status = request.data['status']
        report.save()
        return Response(DailyReportSerializer(report, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='team')
    def team(self, request):
        user = request.user
        if user.role == 'staff':
            return Response({'detail': 'Forbidden'}, status=403)

        qs = DailyReport.objects.select_related('user').all()
        if user.role == 'admin':
            qs = qs.filter(user__role='staff')

        p = request.query_params
        if p.get('month'):
            try:
                y, m = p['month'].split('-')
                qs   = qs.filter(report_date__year=y, report_date__month=m)
            except ValueError:
                pass
        if p.get('week_start'):
            try:
                ws = date.fromisoformat(p['week_start'])
                qs = qs.filter(report_date__range=[ws, ws + timedelta(days=6)])
            except ValueError:
                pass
        if p.get('date'):
            qs = qs.filter(report_date=p['date'])
        if p.get('user_id'):
            qs = qs.filter(user_id=p['user_id'])
        if p.get('report_type'):
            qs = qs.filter(report_type=p['report_type'])

        return Response(DailyReportSerializer(qs[:500], many=True, context={'request': request}).data)

    # ── GET /reports/download/ ── Export weekly or monthly data as xlsx ────────
    @action(detail=False, methods=['get'], url_path='download')
    def download(self, request):
        user       = request.user
        period     = request.query_params.get('period', 'monthly')   # 'weekly' | 'monthly'
        month      = request.query_params.get('month')               # '2026-05'
        week_start = request.query_params.get('week_start')          # '2026-05-26'
        rt_filter  = request.query_params.get('report_type', '')     # specific type or all
        uid        = request.query_params.get('user_id', '')

        qs = DailyReport.objects.select_related('user').all()
        if user.role == 'staff':
            qs = qs.filter(user=user)
        elif user.role == 'admin':
            qs = qs.filter(user__role__in=['staff', 'admin'])

        if uid and user.role in ('admin', 'super_admin'):
            qs = qs.filter(user_id=uid)
        if rt_filter:
            qs = qs.filter(report_type=rt_filter)

        today = date.today()
        if period == 'weekly' and week_start:
            try:
                ws = date.fromisoformat(week_start)
                qs = qs.filter(report_date__range=[ws, ws + timedelta(days=6)])
                period_label = f"Week of {ws.strftime('%d %b %Y')}"
                file_period  = f"week_{week_start}"
            except ValueError:
                return Response({'detail': 'Invalid week_start'}, status=400)
        elif period == 'monthly' and month:
            try:
                y, m = month.split('-')
                qs   = qs.filter(report_date__year=y, report_date__month=m)
                period_label = date(int(y), int(m), 1).strftime('%B %Y')
                file_period  = month
            except ValueError:
                return Response({'detail': 'Invalid month'}, status=400)
        else:
            qs = qs.filter(report_date__year=today.year, report_date__month=today.month)
            period_label = today.strftime('%B %Y')
            file_period  = today.strftime('%Y-%m')

        records = list(qs.order_by('report_type', 'user__first_name', 'report_date'))
        is_admin = user.role in ('admin', 'super_admin')

        wb  = openpyxl.Workbook()
        wb.remove(wb.active)
        hs  = _hdr_style()

        for rt_key, schema in REPORT_SCHEMA.items():
            if rt_filter and rt_filter != rt_key:
                continue
            rt_records = [r for r in records if r.report_type == rt_key]
            ws_sheet   = wb.create_sheet(title=schema['sheet'][:31])
            fields     = schema['fields']

            # Column headers
            headers = (['Date', 'Staff Name'] if is_admin else ['Date']) + [f['label'] for f in fields]
            for ci, hdr in enumerate(headers, 1):
                cell = ws_sheet.cell(row=1, column=ci, value=hdr)
                _apply(cell, hs)
            ws_sheet.row_dimensions[1].height = 22

            if not rt_records:
                ws_sheet.cell(row=2, column=1, value='No data for this period').font = Font(name='Calibri', italic=True, color='9CA3AF', size=10)
            else:
                for ri, rpt in enumerate(rt_records, 2):
                    alt = ri % 2 == 0
                    cs  = _cell_style(alt)
                    if is_admin:
                        row_vals = [
                            rpt.report_date.strftime('%d-%m-%Y'),
                            rpt.user.get_full_name() or rpt.user.username,
                        ] + [rpt.data.get(f['key'], '') or '' for f in fields]
                    else:
                        row_vals = [
                            rpt.report_date.strftime('%d-%m-%Y'),
                        ] + [rpt.data.get(f['key'], '') or '' for f in fields]

                    for ci, val in enumerate(row_vals, 1):
                        cell = ws_sheet.cell(row=ri, column=ci, value=val)
                        _apply(cell, cs)

            # Column widths
            ws_sheet.column_dimensions['A'].width = 14
            if is_admin:
                ws_sheet.column_dimensions['B'].width = 22
            for i in range(2 if not is_admin else 3, len(headers) + 1):
                ws_sheet.column_dimensions[get_column_letter(i)].width = 20

            # Period label in footer
            note_row = (len(rt_records) or 1) + 3
            nc = ws_sheet.cell(row=note_row, column=1,
                               value=f'{schema["label"]} — {period_label}  |  Downloaded {today.strftime("%d %b %Y")}')
            nc.font = Font(name='Calibri', italic=True, color='6B7280', size=9)

            ws_sheet.freeze_panes = 'B2' if not is_admin else 'C2'

        if not wb.sheetnames:
            wb.create_sheet('No Data')

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="reports_{file_period}.xlsx"'
        return resp

    # ── GET /reports/template/ ── Download blank template for a report type ───
    @action(detail=False, methods=['get'], url_path='template')
    def template(self, request):
        rt = request.query_params.get('report_type', '')
        if not rt or rt not in REPORT_SCHEMA:
            return Response({'detail': 'Invalid or missing report_type.'}, status=400)

        schema = REPORT_SCHEMA[rt]
        fields = schema['fields']

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = schema['sheet'][:31]
        hs  = _hdr_style()
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_al = Alignment(horizontal='center', vertical='center')

        headers = ['Date'] + [f['label'] for f in fields]
        for col, hdr in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=hdr)
            _apply(c, hs)
        ws.row_dimensions[1].height = 22

        # Pre-populate last 3 months of dates
        today = date.today()
        dates = []
        for m_offset in range(2, -1, -1):
            y = today.year if today.month > m_offset else today.year - 1
            m = (today.month - m_offset - 1) % 12 + 1
            days_in_month = calendar.monthrange(y, m)[1]
            for d in range(1, days_in_month + 1):
                dates.append(date(y, m, d))

        alt_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        for row_idx, d in enumerate(dates, 2):
            alt = row_idx % 2 == 0
            dc = ws.cell(row=row_idx, column=1, value=d.strftime('%d-%m-%Y'))
            dc.font = Font(name='Calibri', bold=True, size=10)
            dc.border = border; dc.alignment = center_al
            if alt: dc.fill = alt_fill
            for col_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value='')
                cell.font = Font(name='Calibri', size=10)
                cell.border = border; cell.alignment = center_al
                if alt: cell.fill = alt_fill

        ws.column_dimensions['A'].width = 14
        for i in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        note_row = len(dates) + 3
        nc = ws.cell(row=note_row, column=1,
                     value=f'Template: {schema["label"]} — Date format DD-MM-YYYY')
        nc.font = Font(name='Calibri', italic=True, color='6B7280', size=9)

        ws.freeze_panes = 'B2'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        safe = schema['label'].replace(' ', '_').replace('&', 'and').lower()
        resp['Content-Disposition'] = f'attachment; filename="{safe}_template.xlsx"'
        return resp

    # ── POST /reports/import/ ── Upload xlsx to update DB ─────────────────────
    @action(detail=False, methods=['post'], url_path='import')
    def import_data(self, request):
        rt   = LEGACY_TYPE_MAP.get(request.data.get('report_type', ''), request.data.get('report_type', ''))
        file = request.FILES.get('file')

        if not rt or rt not in REPORT_SCHEMA:
            return Response({'detail': 'Invalid or missing report_type.'}, status=400)
        if not file:
            return Response({'detail': 'No file provided.'}, status=400)

        schema = REPORT_SCHEMA[rt]
        fields = schema['fields']

        try:
            df = pd.read_excel(file) if not file.name.lower().endswith('.csv') else pd.read_csv(file)
        except Exception as e:
            return Response({'detail': f'Cannot read file: {e}'}, status=400)

        df.columns = [str(c).strip().lower() for c in df.columns]
        date_col = next((c for c in df.columns if 'date' in c), None)
        if not date_col:
            return Response({'detail': 'No "Date" column found in the file.'}, status=400)

        label_to_key = {f['label'].lower(): f['key'] for f in fields}
        key_set = {f['key'] for f in fields}

        created = updated = skipped = 0
        errors = []

        for idx, row in df.iterrows():
            raw_date = row[date_col]
            if isinstance(raw_date, str): raw_date = raw_date.strip()
            try:
                parsed_date = pd.to_datetime(str(raw_date), dayfirst=True).date()
            except Exception:
                skipped += 1; continue

            data = {}
            for col in df.columns:
                if col == date_col: continue
                # Skip "staff name" column if present
                if col in ('staff name', 'name'): continue
                key = label_to_key.get(col) or (col if col in key_set else None)
                if key:
                    val = row[col]
                    if isinstance(val, float) and math.isnan(val): val = None
                    data[key] = str(val).strip() if val is not None else ''

            if not data: skipped += 1; continue

            try:
                obj, is_new = DailyReport.objects.get_or_create(
                    user=request.user,
                    report_date=parsed_date,
                    report_type=rt,
                    defaults={'data': data, 'status': 'submitted', 'submitted_at': timezone.now()},
                )
                if is_new:
                    created += 1
                else:
                    obj.data.update(data); obj.save(); updated += 1
            except Exception as e:
                errors.append(f'Row {idx + 2}: {e}')

        return Response({
            'message': f'Import complete — {created} created, {updated} updated'
                       + (f', {skipped} skipped' if skipped else ''),
            'created': created, 'updated': updated, 'skipped': skipped,
            'errors': errors[:10],
        })
